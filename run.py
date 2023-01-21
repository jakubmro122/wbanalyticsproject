from flask import Flask, render_template, request
from openpyxl import Workbook, load_workbook
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from io import BytesIO
from PIL import Image
import base64





app = Flask(__name__)





@app.route('/')
def index():
    return render_template('index.html')
        
@app.route('/oneDay/')
def about():

    wb = load_workbook('ZebraneDane.xlsx')
    ws = wb.active
    columName = [ws.cell(row=3,column=i).value for i in range(1,6)]
    data = [ws.cell(row=4,column=i).value for i in range(1,6)]
    machineName = (ws['B1'].value)
    machineStoppage = data[1] - data[2]



    
    return render_template('oneDay.html', columName = columName, data=data, machineName = machineName, machineStoppage = machineStoppage)

@app.route('/oneDay/181')
def oneDay181():
    
     wb = load_workbook('ZebraneDane.xlsx')
     wb.active = wb['181']
     ws = wb.active
     columName = [ws.cell(row=3,column=i).value for i in range(1,6)]
     data = [ws.cell(row=4,column=i).value for i in range(1,6)]
     machineName = (ws['B1'].value)
     machineStoppage = data[1] - data[2]

     return render_template('oneDay.html', columName = columName, data=data, machineName = machineName, machineStoppage = machineStoppage) 

@app.route('/oneDay/230')
def oneDay230():
    
     wb = load_workbook('ZebraneDane.xlsx')
     wb.active = wb['230']
     ws = wb.active
     columName = [ws.cell(row=3,column=i).value for i in range(1,6)]
     data = [ws.cell(row=4,column=i).value for i in range(1,6)]
     machineName = (ws['B1'].value)
     machineStoppage = data[1] - data[2]
    
     return render_template('oneDay.html', columName = columName, data=data, machineName = machineName, machineStoppage = machineStoppage)

@app.route('/oneDay/254')
def oneDay254():
    
     wb = load_workbook('ZebraneDane.xlsx')
     wb.active = wb['254']
     ws = wb.active
     columName = [ws.cell(row=3,column=i).value for i in range(1,6)]
     data = [ws.cell(row=4,column=i).value for i in range(1,6)]
     machineName = (ws['B1'].value)
     machineStoppage = data[1] - data[2]
    
     return render_template('oneDay.html', columName = columName, data=data, machineName = machineName, machineStoppage = machineStoppage)       

@app.route('/oneDay/268')
def oneDay268():
    
     wb = load_workbook('ZebraneDane.xlsx')
     wb.active = wb['268']
     ws = wb.active
     columName = [ws.cell(row=3,column=i).value for i in range(1,6)]
     data = [ws.cell(row=4,column=i).value for i in range(1,6)]
     machineName = (ws['B1'].value)
     machineStoppage = data[1] - data[2]
    
     return render_template('oneDay.html', columName = columName, data=data, machineName = machineName, machineStoppage = machineStoppage)      

@app.route('/oneDay/273')
def oneDay273():
    
     wb = load_workbook('ZebraneDane.xlsx')
     wb.active = wb['273']
     ws = wb.active
     columName = [ws.cell(row=3,column=i).value for i in range(1,6)]
     data = [ws.cell(row=4,column=i).value for i in range(1,6)]
     machineName = (ws['B1'].value)
     machineStoppage = data[1] - data[2]
    
     return render_template('oneDay.html', columName = columName, data=data, machineName = machineName, machineStoppage = machineStoppage)  

@app.route('/oneDay/269', methods=['POST', "GET"])
def oneDay269():
    
     wb = load_workbook('ZebraneDane.xlsx')
     wb.active = wb['269']
     ws = wb.active
     columName = [ws.cell(row=3,column=i).value for i in range(1,6)]
     data = [ws.cell(row=4,column=i).value for i in range(1,6)]
     machineName = (ws['B1'].value)
     machineStoppage = data[1] - data[2]
    
     return render_template('oneDay.html', columName = columName, data=data, machineName = machineName, machineStoppage = machineStoppage)  



@app.route('/sevenDays/')
def seven():
 wb = load_workbook('ZebraneDane.xlsx')
 ws = wb.active
 columName = []


 columName = [ws.cell(row=3,column=i).value for i in range(1,6)]
 sevenDaysData = []
 sevenDaysMachineStoppage = []

 for r in range(0,7):
  data = [ws.cell(row=4+r, column=i).value for i in range(1,6)]
  sevenDaysData.append(data)

  if (data[1] == None):
      sevenDaysMachineStoppage.append(0)
  else:
       machineStoppage = data[1] - data[2]
       sevenDaysMachineStoppage.append(machineStoppage)
       
  machineName = (ws['B1'].value)
  
 chart1 = Image.open("static/generatedDiagrams/181/7/181-7-1.png")
 chart1_file = BytesIO()
 chart1.save(chart1_file, format='PNG')
 chart1_file.seek(0)
 chart1_data = chart1_file.read()
 chart1_b64 = base64.b64encode(chart1_data).decode()

 chart2 = Image.open("static/generatedDiagrams/181/7/181-7-2.png")
 chart2_file = BytesIO()
 chart2.save(chart2_file, format='PNG')
 chart2_file.seek(0)
 chart2_data = chart2_file.read()
 chart2_b64 = base64.b64encode(chart2_data).decode()

 chart3 = Image.open("static/generatedDiagrams/181/7/181-7-3.png")
 chart3_file = BytesIO()
 chart3.save(chart3_file, format='PNG')
 chart3_file.seek(0)
 chart3_data = chart3_file.read()
 chart3_b64 = base64.b64encode(chart3_data).decode()

 chart4 = Image.open("static/generatedDiagrams/181/7/181-7-4.png")
 chart4_file = BytesIO()
 chart4.save(chart4_file, format='PNG')
 chart4_file.seek(0)
 chart4_data = chart4_file.read()
 chart4_b64 = base64.b64encode(chart4_data).decode()

 chart5 = Image.open("static/generatedDiagrams/181/7/181-7-5.png")
 chart5_file = BytesIO()
 chart5.save(chart5_file, format='PNG')
 chart5_file.seek(0)
 chart5_data = chart5_file.read()
 chart5_b64 = base64.b64encode(chart5_data).decode()


 return render_template('sevenDays.html', columName = columName , data=data, machineName = machineName, machineStoppage = machineStoppage, sevenDaysData = sevenDaysData, sevenDaysMachineStoppage = sevenDaysMachineStoppage, chart1=chart1_b64, chart2=chart2_b64, chart3=chart3_b64, chart4=chart4_b64, chart5=chart5_b64)

@app.route('/sevenDays/181', methods=['POST', "GET"])
def sevenDays181():
    
      wb = load_workbook('ZebraneDane.xlsx')
      wb.active = wb['181']
      ws = wb.active
      columName = []

      columName = [ws.cell(row=3,column=i).value for i in range(1,6)]
      sevenDaysData = []
      sevenDaysMachineStoppage = []

      for r in range(0,7):
       data = [ws.cell(row=4+r, column=i).value for i in range(1,6)]
       sevenDaysData.append(data)
       if (data[1] == None):
        sevenDaysMachineStoppage.append(0)
       else:
        machineStoppage = data[1] - data[2]
        sevenDaysMachineStoppage.append(machineStoppage)
      machineName = (ws['B1'].value)

      chart1 = Image.open("static/generatedDiagrams/181/7/181-7-1.png")
      chart1_file = BytesIO()
      chart1.save(chart1_file, format='PNG')
      chart1_file.seek(0)
      chart1_data = chart1_file.read()
      chart1_b64 = base64.b64encode(chart1_data).decode()

      chart2 = Image.open("static/generatedDiagrams/181/7/181-7-2.png")
      chart2_file = BytesIO()
      chart2.save(chart2_file, format='PNG')
      chart2_file.seek(0)
      chart2_data = chart2_file.read()
      chart2_b64 = base64.b64encode(chart2_data).decode()

      chart3 = Image.open("static/generatedDiagrams/181/7/181-7-3.png")
      chart3_file = BytesIO()
      chart3.save(chart3_file, format='PNG')
      chart3_file.seek(0)
      chart3_data = chart3_file.read()
      chart3_b64 = base64.b64encode(chart3_data).decode()

      chart4 = Image.open("static/generatedDiagrams/181/7/181-7-4.png")
      chart4_file = BytesIO()
      chart4.save(chart4_file, format='PNG')
      chart4_file.seek(0)
      chart4_data = chart4_file.read()
      chart4_b64 = base64.b64encode(chart4_data).decode()

      chart5 = Image.open("static/generatedDiagrams/181/7/181-7-5.png")
      chart5_file = BytesIO()
      chart5.save(chart5_file, format='PNG')
      chart5_file.seek(0)
      chart5_data = chart5_file.read()
      chart5_b64 = base64.b64encode(chart5_data).decode()
    
      return render_template('sevenDays.html', columName = columName, data=data, machineName = machineName, machineStoppage = machineStoppage, sevenDaysData = sevenDaysData, sevenDaysMachineStoppage = sevenDaysMachineStoppage, chart1=chart1_b64, chart2=chart2_b64, chart3=chart3_b64, chart4=chart4_b64, chart5=chart5_b64) 

@app.route('/sevenDays/230', methods=['POST', "GET"])
def sevenDays230():
    
      wb = load_workbook('ZebraneDane.xlsx')
      wb.active = wb['230']
      ws = wb.active
      columName = []

      columName = [ws.cell(row=3,column=i).value for i in range(1,6)]
      sevenDaysData = []
      sevenDaysMachineStoppage = []

      for r in range(0,7):
       data = [ws.cell(row=4+r, column=i).value for i in range(1,6)]
       sevenDaysData.append(data)
       if (data[1] == None):
        sevenDaysMachineStoppage.append(0)
       else:
        machineStoppage = data[1] - data[2]
        sevenDaysMachineStoppage.append(machineStoppage)
      machineName = (ws['B1'].value)


      
    
      return render_template('sevenDays.html', columName = columName, data=data, machineName = machineName, machineStoppage = machineStoppage, sevenDaysData = sevenDaysData, sevenDaysMachineStoppage = sevenDaysMachineStoppage)


@app.route('/sevenDays/254', methods=['POST', "GET"])
def sevenDays254():
    
      wb = load_workbook('ZebraneDane.xlsx')
      wb.active = wb['254']
      ws = wb.active
      columName = []

      columName = [ws.cell(row=3,column=i).value for i in range(1,6)]
      sevenDaysData = []
      sevenDaysMachineStoppage = []

      for r in range(0,7):
       data = [ws.cell(row=4+r, column=i).value for i in range(1,6)]
       sevenDaysData.append(data)
       if (data[1] == None):
        sevenDaysMachineStoppage.append(0)
       else:
        machineStoppage = data[1] - data[2]
        sevenDaysMachineStoppage.append(machineStoppage)
      machineName = (ws['B1'].value)
      chart1 = Image.open("static/generatedDiagrams/254/7/254-7-1.png")
      chart1_file = BytesIO()
      chart1.save(chart1_file, format='PNG')
      chart1_file.seek(0)
      chart1_data = chart1_file.read()
      chart1_b64 = base64.b64encode(chart1_data).decode()

      chart2 = Image.open("static/generatedDiagrams/254/7/254-7-2.png")
      chart2_file = BytesIO()
      chart2.save(chart2_file, format='PNG')
      chart2_file.seek(0)
      chart2_data = chart2_file.read()
      chart2_b64 = base64.b64encode(chart2_data).decode()

      chart3 = Image.open("static/generatedDiagrams/254/7/254-7-3.png")
      chart3_file = BytesIO()
      chart3.save(chart3_file, format='PNG')
      chart3_file.seek(0)
      chart3_data = chart3_file.read()
      chart3_b64 = base64.b64encode(chart3_data).decode()

      chart4 = Image.open("static/generatedDiagrams/254/7/254-7-4.png")
      chart4_file = BytesIO()
      chart4.save(chart4_file, format='PNG')
      chart4_file.seek(0)
      chart4_data = chart4_file.read()
      chart4_b64 = base64.b64encode(chart4_data).decode()

      chart5 = Image.open("static/generatedDiagrams/254/7/254-7-5.png")
      chart5_file = BytesIO()
      chart5.save(chart5_file, format='PNG')
      chart5_file.seek(0)
      chart5_data = chart5_file.read()
      chart5_b64 = base64.b64encode(chart5_data).decode()
    
      return render_template('sevenDays.html', columName = columName, data=data, machineName = machineName, machineStoppage = machineStoppage, sevenDaysData = sevenDaysData, sevenDaysMachineStoppage = sevenDaysMachineStoppage, chart1=chart1_b64, chart2=chart2_b64, chart3=chart3_b64, chart4=chart4_b64, chart5=chart5_b64)

@app.route('/sevenDays/268', methods=['POST', "GET"])
def sevenDays268():
    
      wb = load_workbook('ZebraneDane.xlsx')
      wb.active = wb['268']
      ws = wb.active
      columName = []

      columName = [ws.cell(row=3,column=i).value for i in range(1,6)]
      sevenDaysData = []
      sevenDaysMachineStoppage = []

      for r in range(0,7):
       data = [ws.cell(row=4+r, column=i).value for i in range(1,6)]
       sevenDaysData.append(data)
       if (data[1] == None):
        sevenDaysMachineStoppage.append(0)
       else:
        machineStoppage = data[1] - data[2]
        sevenDaysMachineStoppage.append(machineStoppage)
      machineName = (ws['B1'].value)
      chart1 = Image.open("static/generatedDiagrams/268/7/268-7-all-1.png")
      chart1_file = BytesIO()
      chart1.save(chart1_file, format='PNG')
      chart1_file.seek(0)
      chart1_data = chart1_file.read()
      chart1_b64 = base64.b64encode(chart1_data).decode()

      chart2 = Image.open("static/generatedDiagrams/268/7/268-7-all-2.png")
      chart2_file = BytesIO()
      chart2.save(chart2_file, format='PNG')
      chart2_file.seek(0)
      chart2_data = chart2_file.read()
      chart2_b64 = base64.b64encode(chart2_data).decode()

      chart3 = Image.open("static/generatedDiagrams/268/7/268-7-3.png")
      chart3_file = BytesIO()
      chart3.save(chart3_file, format='PNG')
      chart3_file.seek(0)
      chart3_data = chart3_file.read()
      chart3_b64 = base64.b64encode(chart3_data).decode()

      chart4 = Image.open("static/generatedDiagrams/268/7/268-7-all-4.png")
      chart4_file = BytesIO()
      chart4.save(chart4_file, format='PNG')
      chart4_file.seek(0)
      chart4_data = chart4_file.read()
      chart4_b64 = base64.b64encode(chart4_data).decode()

      chart5 = Image.open("static/generatedDiagrams/268/7/268-7-all-5.png")
      chart5_file = BytesIO()
      chart5.save(chart5_file, format='PNG')
      chart5_file.seek(0)
      chart5_data = chart5_file.read()
      chart5_b64 = base64.b64encode(chart5_data).decode()
    
      return render_template('sevenDays.html', columName = columName, data=data, machineName = machineName, machineStoppage = machineStoppage, sevenDaysData = sevenDaysData, sevenDaysMachineStoppage = sevenDaysMachineStoppage, chart1=chart1_b64, chart2=chart2_b64, chart3=chart3_b64, chart4=chart4_b64, chart5=chart5_b64)

@app.route('/sevenDays/273', methods=['POST', "GET"])
def sevenDays273():
    
      wb = load_workbook('ZebraneDane.xlsx')
      wb.active = wb['273']
      ws = wb.active
      columName = []

      columName = [ws.cell(row=3,column=i).value for i in range(1,6)]
      sevenDaysData = []
      sevenDaysMachineStoppage = []

      for r in range(0,7):
       data = [ws.cell(row=4+r, column=i).value for i in range(1,6)]
       sevenDaysData.append(data)
       if (data[1] == None):
        sevenDaysMachineStoppage.append(0)
       else:
        machineStoppage = data[1] - data[2]
        sevenDaysMachineStoppage.append(machineStoppage)
      machineName = (ws['B1'].value)
      chart1 = Image.open("static/generatedDiagrams/273/7/273-7-1.png")
      chart1_file = BytesIO()
      chart1.save(chart1_file, format='PNG')
      chart1_file.seek(0)
      chart1_data = chart1_file.read()
      chart1_b64 = base64.b64encode(chart1_data).decode()

      chart2 = Image.open("static/generatedDiagrams/273/7/273-7-2.png")
      chart2_file = BytesIO()
      chart2.save(chart2_file, format='PNG')
      chart2_file.seek(0)
      chart2_data = chart2_file.read()
      chart2_b64 = base64.b64encode(chart2_data).decode()

      chart3 = Image.open("static/generatedDiagrams/273/7/273-7-3.png")
      chart3_file = BytesIO()
      chart3.save(chart3_file, format='PNG')
      chart3_file.seek(0)
      chart3_data = chart3_file.read()
      chart3_b64 = base64.b64encode(chart3_data).decode()

      chart4 = Image.open("static/generatedDiagrams/273/7/273-7-4.png")
      chart4_file = BytesIO()
      chart4.save(chart4_file, format='PNG')
      chart4_file.seek(0)
      chart4_data = chart4_file.read()
      chart4_b64 = base64.b64encode(chart4_data).decode()

      chart5 = Image.open("static/generatedDiagrams/273/7/273-7-5.png")
      chart5_file = BytesIO()
      chart5.save(chart5_file, format='PNG')
      chart5_file.seek(0)
      chart5_data = chart5_file.read()
      chart5_b64 = base64.b64encode(chart5_data).decode()
    
      return render_template('sevenDays.html', columName = columName, data=data, machineName = machineName, machineStoppage = machineStoppage, sevenDaysData = sevenDaysData, sevenDaysMachineStoppage = sevenDaysMachineStoppage, chart1=chart1_b64, chart2=chart2_b64, chart3=chart3_b64, chart4=chart4_b64, chart5=chart5_b64)

@app.route('/sevenDays/269', methods=['POST', "GET"])
def sevenDays269():
    
      wb = load_workbook('ZebraneDane.xlsx')
      wb.active = wb['269']
      ws = wb.active
      columName = []

      columName = [ws.cell(row=3,column=i).value for i in range(1,6)]
      sevenDaysData = []
      sevenDaysMachineStoppage = []

      for r in range(0,7):
       data = [ws.cell(row=4+r, column=i).value for i in range(1,6)]
       sevenDaysData.append(data)
       if (data[1] == None):
        sevenDaysMachineStoppage.append(0)
       else:
        machineStoppage = data[1] - data[2]
        sevenDaysMachineStoppage.append(machineStoppage)
      machineName = (ws['B1'].value)
      chart1 = Image.open("static/generatedDiagrams/269/7/269-7-1.png")
      chart1_file = BytesIO()
      chart1.save(chart1_file, format='PNG')
      chart1_file.seek(0)
      chart1_data = chart1_file.read()
      chart1_b64 = base64.b64encode(chart1_data).decode()

      chart2 = Image.open("static/generatedDiagrams/269/7/269-7-2.png")
      chart2_file = BytesIO()
      chart2.save(chart2_file, format='PNG')
      chart2_file.seek(0)
      chart2_data = chart2_file.read()
      chart2_b64 = base64.b64encode(chart2_data).decode()

      chart3 = Image.open("static/generatedDiagrams/269/7/269-7-3.png")
      chart3_file = BytesIO()
      chart3.save(chart3_file, format='PNG')
      chart3_file.seek(0)
      chart3_data = chart3_file.read()
      chart3_b64 = base64.b64encode(chart3_data).decode()

      chart4 = Image.open("static/generatedDiagrams/269/7/269-7-4.png")
      chart4_file = BytesIO()
      chart4.save(chart4_file, format='PNG')
      chart4_file.seek(0)
      chart4_data = chart4_file.read()
      chart4_b64 = base64.b64encode(chart4_data).decode()

      chart5 = Image.open("static/generatedDiagrams/269/7/269-7-5.png")
      chart5_file = BytesIO()
      chart5.save(chart5_file, format='PNG')
      chart5_file.seek(0)
      chart5_data = chart5_file.read()
      chart5_b64 = base64.b64encode(chart5_data).decode()
    
      return render_template('sevenDays.html', columName = columName, data=data, machineName = machineName, machineStoppage = machineStoppage, sevenDaysData = sevenDaysData, sevenDaysMachineStoppage = sevenDaysMachineStoppage, chart1=chart1_b64, chart2=chart2_b64, chart3=chart3_b64, chart4=chart4_b64, chart5=chart5_b64)

@app.route('/thirtyDays/')
def thirty():
 wb = load_workbook('ZebraneDane.xlsx')
 ws = wb.active
 columName = []

 columName = [ws.cell(row=3,column=i).value for i in range(1,6)]
 thirtyDaysData = []
 thirtyDaysMachineStoppage = []

 for r in range(0,30):
  data = [ws.cell(row=4+r, column=i).value for i in range(1,6)]
  thirtyDaysData.append(data)

  if (data[1] == None):
        thirtyDaysMachineStoppage.append(0)
  else:
        machineStoppage = data[1] - data[2]
        thirtyDaysMachineStoppage.append(machineStoppage)
  machineName = (ws['B1'].value)

 return render_template('thirtyDays.html',  columName = columName , data=data, machineName = machineName, machineStoppage = machineStoppage, thirtyDaysData = thirtyDaysData, thirtyDaysMachineStoppage=thirtyDaysMachineStoppage)


@app.route('/thirtyDays/181', methods=['POST', "GET"])
def thirtyDays181():
    
     wb = load_workbook('ZebraneDane.xlsx')
     wb.active = wb['181']
     ws = wb.active
     columName = []

     columName = [ws.cell(row=3,column=i).value for i in range(1,6)]
     thirtyDaysData = []
     thirtyDaysMachineStoppage = []

     for r in range(0,30):
      data = [ws.cell(row=4+r, column=i).value for i in range(1,6)]
      thirtyDaysData.append(data)
      machineStoppage = data[1] - data[2]
      thirtyDaysMachineStoppage.append(machineStoppage)

     machineName = (ws['B1'].value)

     chart1 = Image.open("static/generatedDiagrams/181/30/181-30-1.png")
     chart1_file = BytesIO()
     chart1.save(chart1_file, format='PNG')
     chart1_file.seek(0)
     chart1_data = chart1_file.read()
     chart1_b64 = base64.b64encode(chart1_data).decode()
     
     #maszyna 181 30days wykres 2 
     chart2 = Image.open("static/generatedDiagrams/181/30/181-30-2.png")
     chart2_file = BytesIO()
     chart2.save(chart2_file, format='PNG')
     chart2_file.seek(0)
     chart2_data = chart2_file.read()
     chart2_b64 = base64.b64encode(chart2_data).decode()
    #maszyna 181 30days wykres 3 
     chart3 = Image.open("static/generatedDiagrams/181/30/181-30-3.png")
     chart3_file = BytesIO()
     chart3.save(chart3_file, format='PNG')
     chart3_file.seek(0)
     chart3_data = chart3_file.read()
     chart3_b64 = base64.b64encode(chart3_data).decode()

      #maszyna 181 30days wykres 4
     chart4 = Image.open("static/generatedDiagrams/181/30/181-30-4.png")
     chart4_file = BytesIO()
     chart4.save(chart4_file, format='PNG')
     chart4_file.seek(0)
     chart4_data = chart4_file.read()
     chart4_b64 = base64.b64encode(chart4_data).decode()
      #maszyna 181 30days wykres 5
     chart5 = Image.open("static/generatedDiagrams/181/30/181-30-5.png")
     chart5_file = BytesIO()
     chart5.save(chart5_file, format='PNG')
     chart5_file.seek(0)
     chart5_data = chart5_file.read()
     chart5_b64 = base64.b64encode(chart5_data).decode()

     return render_template('thirtyDays.html',  columName = columName , data=data, machineName = machineName, machineStoppage = machineStoppage, thirtyDaysData = thirtyDaysData, thirtyDaysMachineStoppage = thirtyDaysMachineStoppage, chart1=chart1_b64, chart2=chart2_b64, chart3=chart3_b64, chart4=chart4_b64, chart5=chart5_b64)

@app.route('/thirtyDays/230', methods=['POST', "GET"])
def thirtyDays230():
    
     wb = load_workbook('ZebraneDane.xlsx')
     wb.active = wb['230']
     ws = wb.active
     columName = []

     columName = [ws.cell(row=3,column=i).value for i in range(1,6)]
     thirtyDaysData = []
     thirtyDaysMachineStoppage = []

     for r in range(0,30):
      data = [ws.cell(row=4+r, column=i).value for i in range(1,6)]
      thirtyDaysData.append(data)
      if (data[1] == None):
        thirtyDaysMachineStoppage.append(0)
      else:
       machineStoppage = data[1] - data[2]
       thirtyDaysMachineStoppage.append(machineStoppage)
      

     machineName = (ws['B1'].value)

     return render_template('thirtyDays.html',  columName = columName , data=data, machineName = machineName, machineStoppage = machineStoppage, thirtyDaysData = thirtyDaysData, thirtyDaysMachineStoppage = thirtyDaysMachineStoppage)

@app.route('/thirtyDays/254', methods=['POST', "GET"])
def thirtyDays254():
    
     wb = load_workbook('ZebraneDane.xlsx')
     wb.active = wb['254']
     ws = wb.active
     columName = []

     columName = [ws.cell(row=3,column=i).value for i in range(1,6)]
     thirtyDaysData = []
     thirtyDaysMachineStoppage = []

     for r in range(0,30):
      data = [ws.cell(row=4+r, column=i).value for i in range(1,6)]
      thirtyDaysData.append(data)
       
      machineStoppage = data[1] - data[2]
      thirtyDaysMachineStoppage.append(machineStoppage)
      #maszyna 181 30days wykres 1
     machineName = (ws['B1'].value)
     chart1 = Image.open("static/generatedDiagrams/181/30/181-30-1.png")
     chart1_file = BytesIO()
     chart1.save(chart1_file, format='PNG')
     chart1_file.seek(0)
     chart1_data = chart1_file.read()
     chart1_b64 = base64.b64encode(chart1_data).decode()
     
     #maszyna 181 30days wykres 2 
     chart2 = Image.open("static/generatedDiagrams/181/30/181-30-2.png")
     chart2_file = BytesIO()
     chart2.save(chart2_file, format='PNG')
     chart2_file.seek(0)
     chart2_data = chart2_file.read()
     chart2_b64 = base64.b64encode(chart2_data).decode()
    #maszyna 181 30days wykres 3 
     chart3 = Image.open("static/generatedDiagrams/181/30/181-30-3.png")
     chart3_file = BytesIO()
     chart3.save(chart3_file, format='PNG')
     chart3_file.seek(0)
     chart3_data = chart3_file.read()
     chart3_b64 = base64.b64encode(chart3_data).decode()

      #maszyna 181 30days wykres 4
     chart4 = Image.open("static/generatedDiagrams/181/30/181-30-4.png")
     chart4_file = BytesIO()
     chart4.save(chart4_file, format='PNG')
     chart4_file.seek(0)
     chart4_data = chart4_file.read()
     chart4_b64 = base64.b64encode(chart4_data).decode()
      #maszyna 181 30days wykres 5
     chart5 = Image.open("static/generatedDiagrams/181/30/181-30-5.png")
     chart5_file = BytesIO()
     chart5.save(chart5_file, format='PNG')
     chart5_file.seek(0)
     chart5_data = chart5_file.read()
     chart5_b64 = base64.b64encode(chart5_data).decode()

     return render_template('thirtyDays.html',  columName = columName , data=data, machineName = machineName, machineStoppage = machineStoppage, thirtyDaysData = thirtyDaysData, thirtyDaysMachineStoppage = thirtyDaysMachineStoppage, chart1=chart1_b64, chart2=chart2_b64, chart3=chart3_b64, chart4=chart4_b64, chart5=chart5_b64)

@app.route('/thirtyDays/268', methods=['POST', "GET"])
def thirtyDays268():
    
     wb = load_workbook('ZebraneDane.xlsx')
     wb.active = wb['268']
     ws = wb.active
     columName = []

     columName = [ws.cell(row=3,column=i).value for i in range(1,6)]
     thirtyDaysData = []
     thirtyDaysMachineStoppage = []

     for r in range(0,30):
      data = [ws.cell(row=4+r, column=i).value for i in range(1,6)]
      thirtyDaysData.append(data)
      if (data[1] == None):
        thirtyDaysMachineStoppage.append(0)
      else:
       machineStoppage = data[1] - data[2]
       thirtyDaysMachineStoppage.append(machineStoppage)
       #maszyna 268 30days wykres 1
     machineName = (ws['B1'].value)
     chart1 = Image.open("static/generatedDiagrams/268/30/268-30-1.png")
     chart1_file = BytesIO()
     chart1.save(chart1_file, format='PNG')
     chart1_file.seek(0)
     chart1_data = chart1_file.read()
     chart1_b64 = base64.b64encode(chart1_data).decode()
     
     #maszyna 268 30days wykres 2 
     chart2 = Image.open("static/generatedDiagrams/268/30/268-30-2.png")
     chart2_file = BytesIO()
     chart2.save(chart2_file, format='PNG')
     chart2_file.seek(0)
     chart2_data = chart2_file.read()
     chart2_b64 = base64.b64encode(chart2_data).decode()
    #maszyna 268 30days wykres 3 
     chart3 = Image.open("static/generatedDiagrams/268/30/268-30-3.png")
     chart3_file = BytesIO()
     chart3.save(chart3_file, format='PNG')
     chart3_file.seek(0)
     chart3_data = chart3_file.read()
     chart3_b64 = base64.b64encode(chart3_data).decode()

      #maszyna 268 30days wykres 4
     chart4 = Image.open("static/generatedDiagrams/268/30/268-30-4.png")
     chart4_file = BytesIO()
     chart4.save(chart4_file, format='PNG')
     chart4_file.seek(0)
     chart4_data = chart4_file.read()
     chart4_b64 = base64.b64encode(chart4_data).decode()
      #maszyna 268 30days wykres 5
     chart5 = Image.open("static/generatedDiagrams/268/30/268-30-5.png")
     chart5_file = BytesIO()
     chart5.save(chart5_file, format='PNG')
     chart5_file.seek(0)
     chart5_data = chart5_file.read()
     chart5_b64 = base64.b64encode(chart5_data).decode()

     return render_template('thirtyDays.html',  columName = columName , data=data, machineName = machineName, machineStoppage = machineStoppage, thirtyDaysData = thirtyDaysData, thirtyDaysMachineStoppage = thirtyDaysMachineStoppage, chart1=chart1_b64, chart2=chart2_b64, chart3=chart3_b64, chart4=chart4_b64, chart5=chart5_b64)

@app.route('/thirtyDays/273', methods=['POST', "GET"])
def thirtyDays273():
    
     wb = load_workbook('ZebraneDane.xlsx')
     wb.active = wb['273']
     ws = wb.active
     columName = []

     columName = [ws.cell(row=3,column=i).value for i in range(1,6)]
     thirtyDaysData = []
     thirtyDaysMachineStoppage = []

     for r in range(0,30):
      data = [ws.cell(row=4+r, column=i).value for i in range(1,6)]
      thirtyDaysData.append(data)
      machineStoppage = data[1] - data[2]
      thirtyDaysMachineStoppage.append(machineStoppage)
      #maszyna 273 30days wykres 1
     machineName = (ws['B1'].value)
     chart1 = Image.open("static/generatedDiagrams/273/30/273-30-1.png")
     chart1_file = BytesIO()
     chart1.save(chart1_file, format='PNG')
     chart1_file.seek(0)
     chart1_data = chart1_file.read()
     chart1_b64 = base64.b64encode(chart1_data).decode()
     
     #maszyna 273 30days wykres 2 
     chart2 = Image.open("static/generatedDiagrams/273/30/273-30-2.png")
     chart2_file = BytesIO()
     chart2.save(chart2_file, format='PNG')
     chart2_file.seek(0)
     chart2_data = chart2_file.read()
     chart2_b64 = base64.b64encode(chart2_data).decode()
    #maszyna 273 30days wykres 3 
     chart3 = Image.open("static/generatedDiagrams/273/30/273-30-3.png")
     chart3_file = BytesIO()
     chart3.save(chart3_file, format='PNG')
     chart3_file.seek(0)
     chart3_data = chart3_file.read()
     chart3_b64 = base64.b64encode(chart3_data).decode()

      #maszyna 273 30days wykres 4
     chart4 = Image.open("static/generatedDiagrams/273/30/273-30-4.png")
     chart4_file = BytesIO()
     chart4.save(chart4_file, format='PNG')
     chart4_file.seek(0)
     chart4_data = chart4_file.read()
     chart4_b64 = base64.b64encode(chart4_data).decode()
      #maszyna 273 30days wykres 5
     chart5 = Image.open("static/generatedDiagrams/273/30/273-30-5.png")
     chart5_file = BytesIO()
     chart5.save(chart5_file, format='PNG')
     chart5_file.seek(0)
     chart5_data = chart5_file.read()
     chart5_b64 = base64.b64encode(chart5_data).decode()

     return render_template('thirtyDays.html',  columName = columName , data=data, machineName = machineName, machineStoppage = machineStoppage, thirtyDaysData = thirtyDaysData, thirtyDaysMachineStoppage = thirtyDaysMachineStoppage, chart1=chart1_b64, chart2=chart2_b64, chart3=chart3_b64, chart4=chart4_b64, chart5=chart5_b64)

@app.route('/thirtyDays/269', methods=['POST', "GET"])
def thirtyDays269():
    
     wb = load_workbook('ZebraneDane.xlsx')
     wb.active = wb['269']
     ws = wb.active
     columName = []

     columName = [ws.cell(row=3,column=i).value for i in range(1,6)]
     thirtyDaysData = []
     thirtyDaysMachineStoppage = []

     for r in range(0,30):
      data = [ws.cell(row=4+r, column=i).value for i in range(1,6)]
      thirtyDaysData.append(data)
      machineStoppage = data[1] - data[2]
      thirtyDaysMachineStoppage.append(machineStoppage)

     machineName = (ws['B1'].value)
     chart1 = Image.open("static/generatedDiagrams/269/30/269-30-1.png")
     chart1_file = BytesIO()
     chart1.save(chart1_file, format='PNG')
     chart1_file.seek(0)
     chart1_data = chart1_file.read()
     chart1_b64 = base64.b64encode(chart1_data).decode()
     
     #maszyna 181 30days wykres 2 
     chart2 = Image.open("static/generatedDiagrams/269/30/269-30-2.png")
     chart2_file = BytesIO()
     chart2.save(chart2_file, format='PNG')
     chart2_file.seek(0)
     chart2_data = chart2_file.read()
     chart2_b64 = base64.b64encode(chart2_data).decode()
    #maszyna 181 30days wykres 3 
     chart3 = Image.open("static/generatedDiagrams/269/30/269-30-3.png")
     chart3_file = BytesIO()
     chart3.save(chart3_file, format='PNG')
     chart3_file.seek(0)
     chart3_data = chart3_file.read()
     chart3_b64 = base64.b64encode(chart3_data).decode()

      #maszyna 181 30days wykres 4
     chart4 = Image.open("static/generatedDiagrams/269/30/269-30-4.png")
     chart4_file = BytesIO()
     chart4.save(chart4_file, format='PNG')
     chart4_file.seek(0)
     chart4_data = chart4_file.read()
     chart4_b64 = base64.b64encode(chart4_data).decode()
      #maszyna 181 30days wykres 5
     chart5 = Image.open("static/generatedDiagrams/269/30/269-30-5.png")
     chart5_file = BytesIO()
     chart5.save(chart5_file, format='PNG')
     chart5_file.seek(0)
     chart5_data = chart5_file.read()
     chart5_b64 = base64.b64encode(chart5_data).decode()

     return render_template('thirtyDays.html',  columName = columName , data=data, machineName = machineName, machineStoppage = machineStoppage, thirtyDaysData = thirtyDaysData, thirtyDaysMachineStoppage = thirtyDaysMachineStoppage, chart1=chart1_b64, chart2=chart2_b64, chart3=chart3_b64, chart4=chart4_b64, chart5=chart5_b64)

@app.route('/fullData/181', methods = ["POST", "GET"])
def getFullData():
  wb = load_workbook('ZebraneDane.xlsx')
  wb.active = wb['181']
  ws = wb.active
  columName = []

  columName = [ws.cell(row=3,column=i).value for i in range(1,6)]

  fullData = []
  fullMachineStoppage = []

  for r in range(0,57):
      data = [ws.cell(row=4+r, column=i).value for i in range(1,6)]
      fullData.append(data)
      if (data[1] == None):
        fullMachineStoppage.append(0)
      else:
       machineStoppage = data[1] - data[2]
       fullMachineStoppage.append(machineStoppage)
  
  machineName = (ws['B1'].value)
  return render_template('fullData.html', columName = columName, data=data, machineName = machineName, machineStoppage=machineStoppage, fullData = fullData, fullMachineStoppage = fullMachineStoppage)

@app.route('/fullData/230')
def getFullData230():
  wb = load_workbook('ZebraneDane.xlsx')
  wb.active = wb['230']
  ws = wb.active
  columName = []

  columName = [ws.cell(row=3,column=i).value for i in range(1,6)]

  fullData = []
  fullMachineStoppage = []

  for r in range(0,57):
      data = [ws.cell(row=4+r, column=i).value for i in range(1,6)]
      fullData.append(data)
      if (data[1] == None):
        fullMachineStoppage.append(0)
      else:
       machineStoppage = data[1] - data[2]
       fullMachineStoppage.append(machineStoppage)
  
  machineName = (ws['B1'].value)
  return render_template('fullData.html', columName = columName, data=data, machineName = machineName, machineStoppage=machineStoppage, fullData = fullData, fullMachineStoppage = fullMachineStoppage)

@app.route('/fullData/254', methods = ["POST", "GET"])
def getFullData254():
  wb = load_workbook('ZebraneDane.xlsx')
  wb.active = wb['254']
  ws = wb.active
  columName = []

  columName = [ws.cell(row=3,column=i).value for i in range(1,6)]

  fullData = []
  fullMachineStoppage = []

  for r in range(0,57):
      data = [ws.cell(row=4+r, column=i).value for i in range(1,6)]
      fullData.append(data)
      if (data[1] == None):
        fullMachineStoppage.append(0)
      else:
       machineStoppage = data[1] - data[2]
       fullMachineStoppage.append(machineStoppage)
  
  machineName = (ws['B1'].value)
  return render_template('fullData.html', columName = columName, data=data, machineName = machineName, machineStoppage=machineStoppage, fullData = fullData, fullMachineStoppage = fullMachineStoppage)

@app.route('/fullData/268', methods = ["POST", "GET"])
def getFullData268():
  wb = load_workbook('ZebraneDane.xlsx')
  wb.active = wb['268']
  ws = wb.active
  columName = []

  columName = [ws.cell(row=3,column=i).value for i in range(1,6)]

  fullData = []
  fullMachineStoppage = []

  for r in range(0,57):
      data = [ws.cell(row=4+r, column=i).value for i in range(1,6)]
      fullData.append(data)
      if (data[1] == None):
        fullMachineStoppage.append(0)
      else:
       machineStoppage = data[1] - data[2]
       fullMachineStoppage.append(machineStoppage)
  
  machineName = (ws['B1'].value)
  return render_template('fullData.html', columName = columName, data=data, machineName = machineName, machineStoppage=machineStoppage, fullData = fullData, fullMachineStoppage = fullMachineStoppage)

@app.route('/fullData/273', methods = ["POST", "GET"])
def getFullData273():
  wb = load_workbook('ZebraneDane.xlsx')
  wb.active = wb['273']
  ws = wb.active
  columName = []

  columName = [ws.cell(row=3,column=i).value for i in range(1,6)]

  fullData = []
  fullMachineStoppage = []

  for r in range(0,57):
      data = [ws.cell(row=4+r, column=i).value for i in range(1,6)]
     
      fullData.append(data)
       
        
      if (data[1] == None):
        fullMachineStoppage.append("Brak danych")
                 
      else:
       machineStoppage = data[1] - data[2]
       fullMachineStoppage.append(machineStoppage)
  
  machineName = (ws['B1'].value)
  return render_template('fullData.html', columName = columName, data=data, machineName = machineName, machineStoppage=machineStoppage, fullData = fullData, fullMachineStoppage = fullMachineStoppage)

@app.route('/fullData/269', methods = ["POST", "GET"])
def getFullData269():
  wb = load_workbook('ZebraneDane.xlsx')
  wb.active = wb['269']
  ws = wb.active
  columName = []

  columName = [ws.cell(row=3,column=i).value for i in range(1,6)]

  fullData = []
  fullMachineStoppage = []

  for r in range(0,57):
      data = [ws.cell(row=4+r, column=i).value for i in range(1,6)]
      fullData.append(data)
      if (data[1] == None):
        fullMachineStoppage.append(0)
      else:
       machineStoppage = data[1] - data[2]
       fullMachineStoppage.append(machineStoppage)
  
  machineName = (ws['B1'].value)
  return render_template('fullData.html', columName = columName, data=data, machineName = machineName, machineStoppage=machineStoppage, fullData = fullData, fullMachineStoppage = fullMachineStoppage)

if __name__ =='__main__':
    app.run(0.0.0.0:27397)


